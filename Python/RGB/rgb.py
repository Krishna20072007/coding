from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font

end = 255

wb = Workbook()
ws = wb.active
ws.title = "RGB"

for r in range(0, end + 1):
    for g in range(0, end + 1):
        for b in range(0, end + 1):
            ws.append([r, g, b])

ws['A1'] = "R"
ws['B1'] = "G"
ws['C1'] = "B"

# Set background color for each row
for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
    r, g, b = [cell.value for cell in row]
    color = f"{r:02X}{g:02X}{b:02X}"  # Convert RGB to hexadecimal color
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in row:
        cell.fill = fill

# Create separate sheets for each R value
for r in range(end + 1):
    new_ws = wb.create_sheet(title=str(r))
    new_ws['A1'] = "R"
    new_ws['B1'] = "G"
    new_ws['C1'] = "B"
    
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    
    for col in new_ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=3):
        for cell in col:
            cell.font = header_font
            cell.alignment = header_alignment
    
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        if row[0].value == r:
            new_ws.append([cell.value for cell in row])

wb.remove(ws)  # Remove the original "RGB" sheet

wb.save("rgb.xlsx")
