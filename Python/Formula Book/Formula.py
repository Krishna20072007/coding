import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook


def save_formula():
    formula_name = entry_formula_name.get()
    subject = subject_var.get()
    formula = entry_formula.get("1.0", "end-1c")

    try:
        workbook = load_workbook("formulas.xlsx")

        if subject not in workbook.sheetnames:
            workbook.create_sheet(subject)

            new_sheet = workbook[subject]
            new_sheet["A1"] = "Formula Name"
            new_sheet["B1"] = "Formula"

        sheet = workbook[subject]

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if header_row != ("Formula Name", "Formula"):
            sheet.insert_rows(1)
            sheet["A1"] = "Formula Name"
            sheet["B1"] = "Formula"

        sheet.append([formula_name, formula])

        workbook.save("formulas.xlsx")

        clear_fields()
        status_label.config(text="Formula saved successfully.", fg="green")
    except FileNotFoundError:
        status_label.config(text="File not found.", fg="red")


def clear_fields():
    entry_formula_name.delete(0, tk.END)
    entry_formula.delete("1.0", tk.END)


def load_formulas():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        try:
            workbook = load_workbook(file_path)

            formula_list.delete(0, tk.END)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if header_row == ("Formula Name", "Formula"):
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        formula_name, formula = row
                        formula_list.insert(tk.END, formula_name)

            status_label.config(text="Formulas loaded successfully.", fg="green")
        except FileNotFoundError:
            status_label.config(text="File not found.", fg="red")


def display_formula(event):
    selection = formula_list.curselection()
    if selection:
        index = selection[0]
        formula_name = formula_list.get(index)

        workbook = load_workbook("formulas.xlsx")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == formula_name:
                    formula = row[1]

                    entry_formula_name.delete(0, tk.END)
                    entry_formula_name.insert(tk.END, formula_name)

                    entry_formula.delete("1.0", tk.END)
                    entry_formula.insert(tk.END, formula)

                    break


window = tk.Tk()
window.title("Formula Book")

label_formula_name = tk.Label(window, text="Formula Name")
label_formula_name.grid(row=0, column=0, sticky="w")

label_subject = tk.Label(window, text="Subject")
label_subject.grid(row=1, column=0, sticky="w")

label_formula = tk.Label(window, text="Formula")
label_formula.grid(row=2, column=0, sticky="w")

entry_formula_name = tk.Entry(window)
entry_formula_name.grid(row=0, column=1, padx=10, pady=5)

subject_var = tk.StringVar()
subject_dropdown = tk.OptionMenu(window, subject_var, "Maths", "Physics", "Chemistry")
subject_dropdown.grid(row=1, column=1, padx=10, pady=5)

entry_formula = tk.Text(window, height=5, width=40)
entry_formula.grid(row=2, column=1, padx=10, pady=5)

button_save = tk.Button(window, text="Save Formula", command=save_formula)
button_save.grid(row=3, column=0, padx=10, pady=5, sticky="w")

button_clear = tk.Button(window, text="Clear Fields", command=clear_fields)
button_clear.grid(row=3, column=1, padx=10, pady=5, sticky="w")

button_load = tk.Button(window, text="Load Formulas", command=load_formulas)
button_load.grid(row=3, column=1, padx=10, pady=5, sticky="e")

formula_list = tk.Listbox(window, width=30)
formula_list.grid(row=0, column=2, rowspan=4, padx=10, pady=5, sticky="nsew")
formula_list.bind("<<ListboxSelect>>", display_formula)

scrollbar = tk.Scrollbar(window, orient=tk.VERTICAL)
scrollbar.grid(row=0, column=3, rowspan=4, sticky="ns")
formula_list.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=formula_list.yview)

status_label = tk.Label(window, text="", fg="green")
status_label.grid(row=4, column=0, columnspan=3, padx=10, pady=5)

workbook = Workbook()

window.update_idletasks()
window_width = max(window.winfo_width(), 400)
window_height = window.winfo_height()
window.geometry(f"{window_width}x{window_height}")

window.mainloop()
