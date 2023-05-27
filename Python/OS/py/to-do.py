import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

def show_tasks():
    tasks_text.delete(1.0, tk.END)
    for i, task in enumerate(tasks, start=1):
        tasks_text.insert(tk.END, f"{i}. {task}\n")

def add_task():
    task = entry_task.get()
    if task:
        tasks.append(task)
        show_tasks()
        entry_task.delete(0, tk.END)
        messagebox.showinfo("Success", "Task added successfully.")
    else:
        messagebox.showerror("Error", "Please enter a task.")

def mark_complete():
    task_index = int(entry_task_index.get()) - 1
    if 0 <= task_index < len(tasks):
        tasks[task_index] = f"{tasks[task_index]} (completed)"
        show_tasks()
        entry_task_index.delete(0, tk.END)
        messagebox.showinfo("Success", "Task marked as complete.")
    else:
        messagebox.showerror("Error", "Invalid task number.")

def delete_task():
    task_index = int(entry_task_index.get()) - 1
    if 0 <= task_index < len(tasks):
        del tasks[task_index]
        show_tasks()
        entry_task_index.delete(0, tk.END)
        messagebox.showinfo("Success", "Task deleted.")
    else:
        messagebox.showerror("Error", "Invalid task number.")

def save_tasks():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet["A1"] = "Tasks"

    for i, task in enumerate(tasks, start=2):
        sheet[f"A{i}"] = task
    global filename
    filename = "xlsx/tasks.xlsx"
    workbook.save(filename)

def load_tasks():
    try:
        workbook = load_workbook(filename)
        sheet = workbook["Tasks"]
        tasks.clear()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            task = row[0]
            tasks.append(task)

        show_tasks()
    except FileNotFoundError:
        messagebox.showerror("Error", "No tasks file found.")

# Create the main window
window = tk.Tk()
window.title("To-Do List Application")
window.resizable(False, False)
# Create task entry widgets
label_task = tk.Label(window, text="Task:")
label_task.pack()
entry_task = tk.Entry(window, width=50)
entry_task.pack()

# Create buttons for adding, marking as complete, and deleting tasks
button_add = tk.Button(window, text="Add Task", command=add_task)
button_add.pack()

frame_actions = tk.Frame(window)
frame_actions.pack()

button_mark_complete = tk.Button(frame_actions, text="Mark as Complete", command=mark_complete)
button_mark_complete.pack(side=tk.LEFT)

button_delete = tk.Button(frame_actions, text="Delete Task", command=delete_task)
button_delete.pack(side=tk.LEFT)

# Create task index entry and load/save buttons
frame_task_index = tk.Frame(window)
frame_task_index.pack()

label_task_index = tk.Label(frame_task_index, text="Task Index:")
label_task_index.pack(side=tk.LEFT)
entry_task_index = tk.Entry(frame_task_index, width=5)
entry_task_index.pack(side=tk.LEFT)

button_load = tk.Button(window, text="Load Tasks", command=load_tasks)
button_load.pack()

# Create a text widget to display tasks
tasks_text = tk.Text(window, width=50, height=10)
tasks_text.pack()

# Create a list to store tasks
tasks = []

# Load tasks if the file exists
load_tasks()
save_tasks()
# Start the main event loop
window.mainloop()
