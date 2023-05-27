import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from datetime import datetime

# Initialize an empty dictionary to store expenses
expenses = {}

# Set the salary amount
salary = 10000

# Function to add an expense
def add_expense():
    expense_name = category_entry.get()
    amount = float(amount_entry.get())
    month = month_combo.get()
    year = year_combo.get()
    month_year = f"{month} {year}"

    # Create a new sheet for the selected month and year if it doesn't exist
    if month_year not in expenses:
        expenses[month_year] = []
        create_month_sheet(month_year)

    # Append the expense to the expenses dictionary
    expenses[month_year].append((expense_name, amount))

    # Clear the entry fields
    category_entry.delete(0, tk.END)
    amount_entry.delete(0, tk.END)

    # Check if total expenses reach the salary amount
    total_expenses = sum(expense[1] for expense in expenses[month_year])
    if total_expenses >= salary:
        messagebox.showinfo("Alert", "Total expenses have reached or exceeded the salary amount!")

# Function to create a new sheet for a month and year
def create_month_sheet(month_year):
    if month_year not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=month_year)
        sheet["A1"] = "S.No"
        sheet["B1"] = "Expense Name"
        sheet["C1"] = "Amount"
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15

# Function to view expenses and save to Excel
def view_expenses():
    expense_window = tk.Toplevel(root)
    expense_window.title("Expenses")

    for month_year, expense_list in expenses.items():
        month_year_label = tk.Label(expense_window, text=month_year)
        month_year_label.pack()

        if expense_list:
            for i, expense in enumerate(expense_list, start=2):
                expense_label = tk.Label(expense_window, text=f"{expense[0]}: {expense[1]}")
                expense_label.pack()
        else:
            no_expense_label = tk.Label(expense_window, text="No expenses recorded.")
            no_expense_label.pack()

    # Save expenses to Excel
    save_expenses_to_excel()

# Function to save expenses to Excel
def save_expenses_to_excel():
    for month_year, expense_list in expenses.items():
        if month_year not in workbook.sheetnames:
            create_month_sheet(month_year)

        sheet = workbook[month_year]
        existing_rows = sheet.max_row - 1

        for i, expense in enumerate(expense_list, start=1):
            row = existing_rows + i + 1
            sheet.cell(row=row, column=1).value = row - 1
            sheet.cell(row=row, column=2).value = expense[0]
            sheet.cell(row=row, column=3).value = expense[1]

    workbook.save("expenses.xlsx")

# Create the main window
root = tk.Tk()
root.title("Expense Tracker")

# Create category label and entry field
category_label = tk.Label(root, text="Expense Name:")
category_label.grid(row=0, column=0, padx=10, pady=10)
category_entry = tk.Entry(root)
category_entry.grid(row=0, column=1, padx=10, pady=10)

# Create amount label and entry field
amount_label = tk.Label(root, text="Amount:")
amount_label.grid(row=1, column=0, padx=10, pady=10)
amount_entry = tk.Entry(root)
amount_entry.grid(row=1, column=1, padx=10, pady=10)

# Create month selector
month_label = tk.Label(root, text="Month:")
month_label.grid(row=2, column=0, padx=10, pady=10)
months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
month_combo = ttk.Combobox(root, values=months)
month_combo.current(datetime.now().month - 1)
month_combo.grid(row=2, column=1, padx=10, pady=10)

# Create year selector
year_label = tk.Label(root, text="Year:")
year_label.grid(row=3, column=0, padx=10, pady=10)
years = [str(year) for year in range(datetime.now().year - 10, datetime.now().year + 1)]
year_combo = ttk.Combobox(root, values=years)
year_combo.current(10)
year_combo.grid(row=3, column=1, padx=10, pady=10)

# Create buttons for adding expenses and viewing expenses
add_button = tk.Button(root, text="Add Expense", command=add_expense)
add_button.grid(row=4, column=0, padx=10, pady=10, sticky=tk.E+tk.W)
view_button = tk.Button(root, text="View Expenses", command=view_expenses)
view_button.grid(row=4, column=1, padx=10, pady=10, sticky=tk.E+tk.W)

# Create a new workbook
workbook = Workbook()

# Set grid configuration for the main window
root.grid_columnconfigure(1, weight=1)

# Start the GUI event loop
root.mainloop()
