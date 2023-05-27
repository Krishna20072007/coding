import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook, load_workbook


def save_recipe():
    recipe_name = entry_recipe_name.get()
    ingredients = entry_ingredients.get("1.0", "end-1c")
    instructions = entry_instructions.get("1.0", "end-1c")

    try:
        workbook = load_workbook("recipes.xlsx")
        sheet = workbook.active

        # Check if the header row is present in the Excel file
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if header_row != ("Recipe Name", "Ingredients", "Instructions"):
            # If the header row is missing, add it to the Excel file
            sheet.insert_rows(1)
            sheet["A1"] = "Recipe Name"
            sheet["B1"] = "Ingredients"
            sheet["C1"] = "Instructions"

        # Create a new row in the Excel workbook and add the recipe details
        sheet.append([recipe_name, ingredients, instructions])

        # Save the workbook
        workbook.save("xlsx/recipes.xlsx")

        clear_fields()
        status_label.config(text="Recipe saved successfully.", fg="green")
    except FileNotFoundError:
        status_label.config(text="File not found.", fg="red")


def clear_fields():
    entry_recipe_name.delete(0, tk.END)
    entry_ingredients.delete("1.0", tk.END)
    entry_instructions.delete("1.0", tk.END)


def load_recipes():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Clear the existing recipe list
            recipe_list.delete(0, tk.END)

            # Read the header row
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            if header_row == ("Recipe Name", "Ingredients", "Instructions"):
                # Iterate through the rows and add each recipe to the list
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    recipe_name, ingredients, instructions = row
                    recipe_list.insert(tk.END, recipe_name)

                status_label.config(text="Recipes loaded successfully.", fg="green")
            else:
                status_label.config(text="Invalid file format.", fg="red")
        except FileNotFoundError:
            status_label.config(text="File not found.", fg="red")


def display_recipe(event):
    selection = recipe_list.curselection()
    if selection:
        index = selection[0]
        recipe_name = recipe_list.get(index)

        workbook = load_workbook("recipes.xlsx")
        sheet = workbook.active

        # Find the recipe row based on the selected recipe name
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == recipe_name:
                ingredients = row[1]
                instructions = row[2]

                entry_recipe_name.delete(0, tk.END)
                entry_recipe_name.insert(tk.END, recipe_name)

                entry_ingredients.delete("1.0", tk.END)
                entry_ingredients.insert(tk.END, ingredients)

                entry_instructions.delete("1.0", tk.END)
                entry_instructions.insert(tk.END, instructions)

                break


# Create the main window
window = tk.Tk()
window.title("Cookbook")
window.geometry("600x400")
window.resizable(False, False)

# Create labels
label_recipe_name = tk.Label(window, text="Recipe Name")
label_recipe_name.grid(row=0, column=0, sticky="w")

# Create entry fields
entry_recipe_name = tk.Entry(window)
entry_recipe_name.grid(row=0, column=1, padx=10, pady=5)

label_ingredients = tk.Label(window, text="Ingredients")
label_ingredients.grid(row=1, column=0, sticky="w")

entry_ingredients = tk.Text(window, height=5, width=40)
entry_ingredients.grid(row=1, column=1, padx=10, pady=5)

label_instructions = tk.Label(window, text="Instructions")
label_instructions.grid(row=2, column=0, sticky="w")

entry_instructions = tk.Text(window, height=10, width=40)
entry_instructions.grid(row=2, column=1, padx=10, pady=5)

# Create buttons
button_save = tk.Button(window, text="Save Recipe", command=save_recipe)
button_save.grid(row=3, column=0, padx=10, pady=5, sticky="w")

button_clear = tk.Button(window, text="Clear Fields", command=clear_fields)
button_clear.grid(row=3, column=1, padx=10, pady=5, sticky="w")

button_load = tk.Button(window, text="Load Recipes", command=load_recipes)
button_load.grid(row=3, column=1, padx=10, pady=5, sticky="e")

# Create recipe list
recipe_list = tk.Listbox(window, width=30)
recipe_list.grid(row=0, column=2, rowspan=4, padx=10, pady=5, sticky="nsew")
recipe_list.bind("<<ListboxSelect>>", display_recipe)

# Create a scrollbar for the recipe list
scrollbar = tk.Scrollbar(window, orient=tk.VERTICAL)
scrollbar.grid(row=0, column=3, rowspan=4, sticky="ns")
recipe_list.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=recipe_list.yview)

# Create status label
status_label = tk.Label(window, text="", fg="green")
status_label.grid(row=4, column=0, columnspan=3, padx=10, pady=5)

# Create a workbook object
workbook = Workbook()

# Start the GUI event loop
window.mainloop()
